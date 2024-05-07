import os.path
import glob
from collections.abc import Iterable
import win32com.client

import openpyxl
import requests
from bs4 import BeautifulSoup

START_ROW_VEDOM = 5
END_ROW_VEDOM = 38

abspath = os.getcwd()
DIRECTORY = os.path.join(abspath, 'файлы')


def parce_vedom():
    vedom = glob.glob(os.path.join(DIRECTORY, '*.xlsx'))[0]
    file = openpyxl.load_workbook(vedom, data_only=True)
    sheet = file["ВОР"]

    name_work = [cell.value.strip() for cell in sheet['B'][START_ROW_VEDOM - 1:END_ROW_VEDOM] if cell.value]
    GSN = [cell.value.strip().lower().replace('гэсн', '').strip() for cell in sheet['E'][START_ROW_VEDOM - 1:END_ROW_VEDOM]
           if cell.value]
    units_measurement = [cell.value for cell in sheet['C'][START_ROW_VEDOM - 1:END_ROW_VEDOM] if cell.value]
    volume = [cell.value for cell in sheet['D'][START_ROW_VEDOM - 1:END_ROW_VEDOM] if cell.value]

    file.close()

    return [{work: {'gsn': gsn, 'volume': vol, 'units_measurement': un}}
            for work, gsn, vol, un in zip(name_work, GSN, volume, units_measurement)]


def parce_defsmeta(gsn):
    resources = {g: [] for g in gsn}
    temp_url = r'https://www.defsmeta.com/rgsn/gsn_{}/giesn-{}.php'

    for code in gsn:
        main_code = code.split('-')[0]
        url = temp_url.format(main_code, code)
        page = requests.get(url)
        soup = BeautifulSoup(page.text, 'lxml')

        for label, type_res in zip(
                ('РАСХОД МАТЕРИАЛОВ', 'ЭКСПЛУАТАЦИЯ МАШИН И МЕХАНИЗМОВ', 'ТРУДОЗАТРАТЫ'),
                ('material', 'machine', 'people')
        ):
            start_element = soup.find('p', string=label)
            if start_element:
                elements_between = []
                current_element = start_element
                for _ in range(3):
                    elements_between.append(current_element)
                    current_element = current_element.find_next()
                table = elements_between[2]
                rows = table.find_all('tr')
                if label == 'ТРУДОЗАТРАТЫ':
                    rows = rows[1:-len(rows) // 2]
                else:
                    rows = rows[1:-1]

                for material in rows:
                    material = material.find_all('td')
                    if label == 'ТРУДОЗАТРАТЫ':
                        material = material[1:]
                        consumption = float(material[2].text.replace(',', '.'))
                        unit = material[1].text
                        name = material[0].text
                    else:
                        material = material[1:-1]
                        consumption = float(material[3].text.replace(',', '.'))
                        unit = material[2].text
                        name = material[1].text

                    resources[code].append({
                        'name': name,
                        'consumption': consumption,
                        'type': type_res,
                        'unit': unit.replace('\xa0', '')
                    })
                    if resources[code][-1]['unit'] == 'м':
                        resources[code][-1]['unit'] = 'метры'

    return resources


class MSProject:
    def __init__(self, file):
        self.mpp = win32com.client.Dispatch("MSProject.Application")
        self.mpp.FileOpen(file)
        self.project = self.mpp.ActiveProject

    def delete_resources(self, resource_name: Iterable):
        for r in self.project.Resources:
            if r.name in resource_name:
                r.Delete()

    def delete_all_resources(self):
        self.delete_resources(self.get_name_of_resources())

    def add_resources(self, resource_name: Iterable):
        for name in resource_name:
            self.project.Resources.Add(Name=name)

    def get_name_of_resources(self):
        return [r.name for r in self.project.Resources]

    def get_resources_name_id(self):
        return {t.Name: t.Id for t in self.project.Resources}

    def get_tasks_id_name(self):
        return {t.Id: t.Name for t in self.project.Tasks}

    def get_resources_object(self):
        return self.project.Resources

    def get_tasks_object(self):
        return self.project.Tasks

    def _save_file(self):
        self.mpp.FileSave()

    def _close_file(self):
        self.mpp.Quit()

    def close(self):
        self._save_file()
        self._close_file()


from openpyxl import Workbook


def main():
    project = MSProject(glob.glob(os.path.join(DIRECTORY, '*.mpp'))[0])
    vedom = parce_vedom()[:]

    count_vedom = len(vedom)
    gsns = []
    for w in vedom:
        for g in w.values():
            gsns.append(g['gsn'])

    resources_defsmeta = parce_defsmeta(set(gsns))

    # Добавление всех ресурсов в проект
    set_resources = set()
    for lst in resources_defsmeta.values():
        for r in lst:
            set_resources.add(r['name'])

    project.add_resources(set_resources)

    # Добавление ресурсов в задачу
    resources_obj = project.get_resources_object()
    tasks_obj = project.get_tasks_object()
    list_tasks_obj = list(tasks_obj)
    resources = project.get_resources_name_id()

    for ind, task_vedom in enumerate(vedom, start=1):
        print(f'{ind}/{count_vedom}')
        for task_obj in list_tasks_obj:
            name_task_vedom = list(task_vedom.keys())[0].strip()
            if task_obj.Name.strip() == name_task_vedom:
                gsn = task_vedom[name_task_vedom]['gsn']
                gsn_def = resources_defsmeta[gsn]
                if not list(task_obj.Assignments):
                    for res_name in gsn_def:
                        res_id = resources[res_name['name']]
                        res_obj = resources_obj.Item(res_id)

                        if res_name['type'] == 'material':
                            res_obj.Type = 1
                            res_obj.MaterialLabel = res_name['unit']
                        elif res_name['type'] == 'machine':
                            res_obj.Group = "Машины"

                        elif res_name['type'] == 'people':
                            res_obj.Group = "Рабочие"

                        try:
                            task_obj.FixedDuration = True
                            t = task_obj.Assignments.Add(task_obj.ID, res_id)
                            if res_name['type'] == 'material':
                                t.Units = res_name['consumption'] * task_vedom[name_task_vedom]['volume']
                            else:
                                t.Work = res_name['consumption'] * task_vedom[name_task_vedom]['volume'] * 60


                        except Exception as e:
                            print(e)
                    project._save_file()
                    break
    project.close()
    print('MSProject сформирован')
    print('Создание сводного файла по ресурсам')
    # Создание сводного файла по ресурсам
    new_vedom = [['Наименование работ', 'Объем', ' ', 'ГЭСН', 'Материалы', 'ед. изм.', 'расход', ' '],
                 [' ', 'Ед. изм.', 'Кол-во', ' ', ' ', ' ', 'Норм. на един работ', 'Кол-во'],
                 ]

    summ_vedom = {}

    for ind, task_vedom in enumerate(vedom, start=1):
        new_vedom.append([])
        print(f'{ind}/{count_vedom}')
        name_task_vedom = list(task_vedom.keys())[0].strip()
        work = task_vedom[name_task_vedom]
        gsn = work['gsn']
        gsn_def = resources_defsmeta[gsn]
        new_vedom[-1].append(name_task_vedom)
        new_vedom[-1].append(work['units_measurement'])
        new_vedom[-1].append(work['volume'])
        new_vedom[-1].append(gsn)
        for res_name in gsn_def:
            if res_name['type'] == 'material':
                new_vedom.append([None, None, None, None])
                new_vedom[-1].append(res_name['name'])
                new_vedom[-1].append(res_name['unit'])
                new_vedom[-1].append(res_name['consumption'])
                new_vedom[-1].append(res_name['consumption'] * task_vedom[name_task_vedom]['volume'])
                if summ_vedom.get(res_name['name']) is None:
                    summ_vedom[res_name['name']] = res_name['consumption'] * task_vedom[name_task_vedom]['volume']
                else:
                    summ_vedom[res_name['name']] += res_name['consumption'] * task_vedom[name_task_vedom]['volume']

    wb = Workbook()
    ws1 = wb.create_sheet("Ведомость ресурсов")
    for row in new_vedom:
        ws1.append(row)

    ws1 = wb.create_sheet("Сводная таблица")
    for row in [[k, v] for k, v in summ_vedom.items()]:
        ws1.append(row)

    del wb['Sheet']
    wb.save('Материалы.xlsx')
    #os.startfile('Материалы.xlsx')


if __name__ == '__main__':
    main()